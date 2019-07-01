import openpyxl
import pandas as pd
import time
import numpy as np
import os
import datetime
from openpyxl.chart import (Reference, BarChart, LineChart, RadarChart, AreaChart, DoughnutChart, ProjectedPieChart)
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import numbers
from openpyxl.chart.series import DataPoint


class AutoGenerate(object):
    # 初始化部分
    def __init__(self):
        # df1是数据源台账，df是每个数据可能的返回值，wb是所有程序所有列表的汇总表
        # df2是数据源2财务报表
        # month = input('请输入需要制作报表的月份\n如果不按格式填写会导致程序运行出错\n格式为阿拉伯数字\n例:1\n')
        # month = int(month)
        month = 3  # 真的弄得时候替换掉这个
        month1 = str(month) + '月'
        print('开始读取数据\n')
        g1 = time.process_time()
        this_year = time.localtime(time.time()).tm_year  # this_year指的是2019
        if month == 12:
            this_year -= 1
        this_year1 = this_year - 2000  # this_year指的是19
        this_year2 = this_year1 - 1  # this_year指的是18
        self.year1 = str(this_year1)
        self.year2 = str(this_year2)
        df2 = pd.read_excel('数据源2财务报表.xlsx', sheet_name=(self.year1 + '统计报表明细'), index_col='项目')
        self.df2 = df2.drop(columns=[1, '数据口径', '来源', '合计'])
        df2ly = pd.read_excel('数据源2财务报表（去年）.xlsx', sheet_name=(self.year2 + '统计报表明细'), index_col='项目')
        self.df2ly = df2ly.drop(columns=[1, '数据口径', '来源', '合计'])
        df3 = pd.read_excel('数据源3维保站点.xlsx')
        df3 = df3[['机构名称', '所属分公司']]
        df3.rename(columns={'所属分公司': '分公司'}, inplace=True)
        df3.dropna(axis=0, how='any', inplace=True)
        df3['分公司'] = df3['分公司'].apply(lambda x: x.replace('分公司', ''))
        df3['机构名称'].replace(regex={r'^.*维修站$': '维保站', r'^.*维修点$': '维保点', r'^.*维修中心$': '维保中心'}, inplace=True)
        self.df3 = df3[df3['机构名称'].isin(['维保中心', '维保站', '维保点'])]
        df4 = pd.read_excel('数据源4维保人员.xlsx')
        df4.rename(columns={'岗位': '分公司'}, inplace=True)
        df4.fillna(0, inplace=True)
        try:
            df4 = df4[~df4['分公司'].isin(['总计'])]
            del df4['总计']
        except:
            pass
        df4['分公司'] = df4['分公司'].apply(lambda x: x.replace('分公司', ''))
        self.df4 = df4.copy()
        df5 = pd.read_excel('数据源5远程监视.xlsx')
        df5 = df5[['分公司', '电梯在线状态']]
        df5['分公司'].fillna('无信息', inplace=True)
        df5['分公司'] = df5['分公司'].apply(lambda x: x.replace('分公司', ''))
        self.df5 = df5.copy()
        df6 = pd.read_excel('数据源6急修工单.xlsx')
        df6 = df6[['受信方式', '合同类型', '分公司', '故障性质', '非产品故障']]
        df6 = df6[df6['合同类型'].isin(['自保养合同'])]
        df6['分公司'] = df6['分公司'].apply(lambda x: x.replace('分公司', ''))
        self.df6 = df6.copy()
        wb = openpyxl.load_workbook('自动化程序需要的信息.xlsx')
        ws1 = wb['分公司对应表']
        ws2 = wb['省级分公司排序']
        ws3 = wb['梯型对照表']
        ws4 = wb['楼宇类型']
        self.replace_country = {}  # 这个是用来替换的
        self.sort_province = []  # 这个是用来省级分公司排序的
        self.province2company = {}
        self.replace_province = {}
        sort_company = []  # 这个是用来所有分公司排序的
        elevator_type = {}  # 这个是梯型对应表
        building_type = {}
        for i in range(2, ws1.max_row + 1):
            if ws1.cell(i, 3).value != ws1.cell(i, 1).value:
                self.replace_province[ws1.cell(i, 3).value] = ws1.cell(i, 1).value
        for i in range(2, ws1.max_row + 1):
            if ws1.cell(i, 3).value != ws1.cell(i, 2).value:
                self.replace_country[ws1.cell(i, 3).value] = ws1.cell(i, 2).value
        for i in range(2, ws2.max_row + 1):
            self.sort_province.append(ws2.cell(i, 1).value)
        for i in range(2, ws1.max_row + 1):
            sort_company.append(ws1.cell(i, 3).value)
        for i in range(2, ws1.max_row + 1):
            self.province2company.setdefault(ws1.cell(i, 2).value, []).append(ws1.cell(i, 3).value)
        for i in range(2, ws3.max_row + 1):
            elevator_type[ws3.cell(i, 2).value] = ws3.cell(i, 1).value
        for i in range(2, ws4.max_row + 1):
            building_type[ws4.cell(i, 1).value] = ws4.cell(i, 2).value
        df7 = pd.read_excel('数据源7用户满意度.xlsx')
        df7 = df7[df7['分公司'].str.contains('分公司')]
        df7['保养用户满意度评分'] = df7[['服务意识态度良好', '与用户的工作配合及时有效', '保养人员品行廉洁', '保养现场规范作业', '急修服务响应']].mean(axis=1)
        df7 = df7[['分公司', '保养用户满意度评分', '备件供应服务']]
        df7['分公司'] = df7['分公司'].apply(lambda x: x.replace('分公司', ''))
        self.df7 = df7.copy()
        df8 = pd.read_excel('数据源8移动终端.xlsx')
        df8 = df8[['批次号', '综合执行工单数', '工单总数', '所属分公司']]
        dt = datetime.date(this_year, month, 1)
        df8 = df8[df8['批次号'].isin([dt])]
        del df8['批次号']
        df8.rename(columns={'所属分公司': '分公司'}, inplace=True)
        df8['分公司'] = df8['分公司'].apply(lambda x: x.replace('分公司', ''))
        self.df8 = df8.copy()
        df9 = pd.read_excel('数据源9市场投放量.xlsx', header=1)
        total = '累计到' + str(this_year - 1 - 2000) + '编号投放量'
        df9 = df9[['分公司', total]]
        df9.rename(columns={total: '总投放量'}, inplace=True)
        df9.dropna(axis=0, how='any', inplace=True)
        df9 = df9[df9['分公司'].str.contains('分公司 小计')]
        df9['分公司'] = df9['分公司'].apply(lambda x: x.replace('分公司 小计', ''))
        self.df9 = df9.copy()
        # df10是总保养台数指标
        df10 = pd.read_excel('数据源10指标.xlsx', sheet_name='表6 - 总保养台数分解', header=[1, 2])
        year2 = str(this_year - 1) + '年'

        df10 = df10[[('分公司2', '分公司2'), (month1, '累计'), (month1, '净增'), ('12月', year2)]]
        df10 = df10.droplevel(0, axis=1)
        df10.rename(columns={
            '分公司2': '分公司',
            year2: '去年保养台数'
        }, inplace=True)
        df10.dropna(axis=0, how='any', inplace=True)
        df10 = df10[df10['分公司'].str.contains('分公司')]
        df10['分公司'] = df10['分公司'].apply(lambda x: x.replace('分公司', ''))
        self.df10 = df10.copy()

        # df10a是保养收入指标
        df10a = pd.read_excel('数据源10指标.xlsx', sheet_name='表15 - 保养收入毛利分解', header=[1, 2])
        df10a = df10a[[('分公司2', '分公司2'), (month1, '收入')]]
        df10a = df10a.droplevel(0, axis=1)
        df10a.rename(columns={
            '分公司2': '分公司',
        }, inplace=True)
        df10a.dropna(axis=0, how='any', inplace=True)
        df10a = df10a[df10a['分公司'].str.contains('分公司')]
        df10a['分公司'] = df10a['分公司'].apply(lambda x: x.replace('分公司', ''))
        self.df10a = df10a.copy()

        # df10b是备件收入指标
        df10b = pd.read_excel('数据源10指标.xlsx', sheet_name='表20 - 备件收入毛利分解', header=[1, 2])
        df10b = df10b[[('分公司2', '分公司2'), (month1, '收入')]]
        df10b = df10b.droplevel(0, axis=1)
        df10b.rename(columns={
            '分公司2': '分公司',
        }, inplace=True)
        df10b.dropna(axis=0, how='any', inplace=True)
        df10b = df10b[df10b['分公司'].str.contains('分公司')]
        df10b['分公司'] = df10b['分公司'].apply(lambda x: x.replace('分公司', ''))
        self.df10b = df10b.copy()

        df1 = pd.read_excel('数据源1台账.xlsx')
        df1['楼宇用途'].replace(building_type, inplace=True)
        df1['产品型号'].replace(elevator_type, inplace=True)
        df1['分公司'] = df1['分公司'].astype('category')
        df1['分公司'].cat.set_categories(sort_company, inplace=True)
        df1.sort_values(by=['分公司'], ascending=True, inplace=True)  # 怀疑这一步没必要
        df1['分公司'] = df1['分公司'].astype(object)
        self.df1 = df1.copy()
        g2 = time.process_time()
        g = g2 - g1
        print('数据读取完成,耗时%.2f秒\n' % g)
        os.makedirs('.\\各分公司季度报表', exist_ok=True)

    # 一、辅助函数（开始）

    def list_generate(self):  # 这个用来自动生成省级分公司列表
        content = []
        df = self.df1['分公司'].drop_duplicates()
        df.replace(self.replace_country, inplace=True)
        df = df.drop_duplicates()
        for each in df:
            content.append(each)
        return content

    @staticmethod
    def convert_int(value):  # 这个是除了空值外，都变成整数的函数
        if np.isnan(value):
            if np.isnan(value):
                return '数据缺失'
            else:
                return np.int(value)

    def convert_province(self, data):  # 这个是通过省级分公司得到下属分公司列表
        company = self.province2company[data]
        return company

    @staticmethod
    def value_get(ws, param1, param2='合计'):  # 这个程序是获得合计行中特定列名的值的
        for each_col in ws[1]:  # 获得列坐标
            if each_col.value == param1:
                for each_row in ws['A']:  # 获得行坐标
                    if each_row.value == param2:
                        return ws.cell(each_row.row, each_col.column).value
        return 0

    def company_convert(self, df, default):  # 按照全国计算或者只保留特定分公司
        if default == '全国':
            df.replace(self.replace_province, inplace=True)
        else:
            company = self.convert_province(default)
            df = df[df['分公司'].isin(company)]
        return df

    @staticmethod
    def format_change(ws, param1, nf):  # 第一个参数是列名，第二个参数是工作表，第三个参数是内置自定义格式的名称
        # 试一下不返回ws的话还可以修改成功吗，成功
        for each_line in ws:
            for each_cell in each_line:
                if each_cell.value == param1:
                    col = openpyxl.cell.cell.get_column_letter(each_cell.column)
                    for each in ws[col]:
                        each.number_format = nf
                    return None

    @staticmethod
    def years(x):
        a = time.localtime(time.time()).tm_year
        if x < 80:
            x += 2000
        else:
            x += 1900
        b = a - x
        if b <= 5:
            return '0-5年'
        elif 5 < b <= 10:
            return '5-10年'
        elif 10 < b <= 15:
            return '10-15年'
        else:
            return '15年以上'

    @staticmethod
    def barchart_function(width1, data1, cats1, title1, ytitle1, type1='col', style1=10, gapwidth1=50,
                          grouping1='standard'):
        chart = BarChart()
        chart.type = type1
        chart.style = style1
        chart.width = width1
        chart.grouping = grouping1
        if chart.grouping == 'stacked':
            chart.overlap = 100
        chart.gapWidth = gapwidth1
        chart.y_axis.majorGridlines = None
        chart.y_axis.title = ytitle1
        chart.title = title1
        chart.add_data(data1, titles_from_data=True)
        chart.set_categories(cats1)
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.position = 'bestFit'
        return chart

    @staticmethod
    def radarchart_function(data1, cats1, title1, type1='standard', style1=10):
        chart = RadarChart()
        chart.type = type1
        chart.style = style1
        chart.title = title1
        chart.add_data(data1, titles_from_data=True)
        chart.set_categories(cats1)
        chart.y_axis.delete = True
        chart.y_axis.majorGridlines = None
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.position = 'bestFit'
        return chart

    @staticmethod
    def areachart_function(width1, data1, cats1, title1, grouping1='standard', style1=10, gapwidth1=50):
        chart = AreaChart()
        chart.style = style1
        chart.grouping = grouping1
        chart.width = width1
        chart.gapWidth = gapwidth1
        chart.y_axis.majorGridlines = None
        chart.y_axis.delete = True
        chart.title = title1
        chart.add_data(data1, titles_from_data=True)
        chart.set_categories(cats1)
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.position = 'bestFit'
        return chart

    @staticmethod
    def linechart_function(data, ytitle1):
        chart = LineChart()
        chart.y_axis.majorGridlines = None
        chart.y_axis.axId = 200
        chart.y_axis.title = ytitle1
        chart.add_data(data, titles_from_data=True)
        for each in chart.series:
            each.marker.symbol = 'auto'
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.position = 'bestFit'
        chart.smooth = True
        return chart

    # 一、辅助函数（结束）

    # 二、pandas计算函数开始
    def maintenance(self, default):  # 这就是保养实物量的计算函数
        # df是保养实物量所用的表
        df = self.df1[['数量', '分公司', '类型']].copy()
        dfa = self.df9.copy()
        dfb = self.df10.copy()
        df = self.company_convert(df, default)
        dfa = self.company_convert(dfa, default)
        dfb = self.company_convert(dfb, default)
        left = df['数量'].groupby(df['分公司']).sum()
        right = df.groupby(['分公司', '类型'], as_index=False)['数量'].sum()
        right = right.pivot(index='分公司', columns='类型', values='数量')
        df = pd.merge(left, right, on='分公司')
        df.rename(columns={
            '数量': '总保养台数',
            'ZWV': '自保养台数',
            'ZWV2': '三方合作',
            'ZWV5': '三方合同'
        }, inplace=True)
        # 这个是调整顺序用的

        dfa = dfa.groupby(dfa['分公司']).sum()
        dfb = dfb.groupby(dfb['分公司']).sum()
        df['月度累计计划'] = dfb['累计']
        df['月度净增计划'] = dfb['净增']
        df['去年保养台数'] = dfb['去年保养台数']
        df['本月净增'] = df['总保养台数'] - df['去年保养台数']
        df['总投放量'] = dfa['总投放量']
        df.loc['合计'] = df.sum()
        df['保养市场占有率'] = df['总保养台数'] / df['总投放量']
        df['总保养月度计划完成率'] = df['总保养台数'] / df['月度累计计划']
        df['总保养净增计划完成率'] = df['本月净增'] / df['月度净增计划']
        dfc = df[['总保养台数']].copy()
        try:
            dfc['自保养台数'] = df['自保养台数'].copy()
        except:
            pass
        try:
            dfc['三方合同'] = df['三方合同'].copy()
        except:  # 而且有些公司实际上并没有三方合同
            pass
        try:
            dfc['三方合作'] = df['三方合作'].copy()
        except:  # 而且有些公司实际上并没有三方合作
            pass
        dfc[['保养市场占有率', '总保养月度计划完成率', '总保养净增计划完成率']] = df[['保养市场占有率', '总保养月度计划完成率', '总保养净增计划完成率']].copy()
        return dfc

    def maintenance_value(self, default):  # 这个是保养收入的表
        df = self.df2.copy()
        df = df.loc[['2.保养', ' 2.2销售承诺保养（安装2）']]
        df = df.iloc[[0, 2]]
        df = df.T
        df.rename(columns={
            '2.保养': (self.year1 + '年累计保养收入'),
            ' 2.2销售承诺保养（安装2）': (self.year1 + '年累计免费保养收入')
        }, inplace=True)
        df.index.name = '分公司'
        df.reset_index(inplace=True)
        df = self.company_convert(df, default)  # 公司已经过滤
        df = df[[self.year1 + '年累计保养收入', self.year1 + '年累计免费保养收入']].groupby(df['分公司']).sum()

        dfly = self.df2ly.copy()  # dfly是去年指标
        dfly = dfly.loc[['2.保养']]
        dfly = dfly.iloc[0].to_frame()
        dfly.rename(columns={
            '2.保养': self.year2 + '年同期累计收入'
        }, inplace=True
        )
        dfly.index.name = '分公司'
        dfly.reset_index(inplace=True)
        dfly = self.company_convert(dfly, default)
        dfly = dfly[self.year2 + '年同期累计收入'].groupby(dfly['分公司']).sum().to_frame()
        df[self.year2 + '年同期累计收入'] = dfly
        df = df / 10000  # 指标都是万元，得把其他数据弄下来

        df10a = self.df10a.copy()
        df10a = self.company_convert(df10a, default)
        df10a = df10a['收入'].groupby(df10a['分公司']).sum().to_frame()
        df['指标'] = df10a
        df.loc['合计'] = df.sum()
        df['月度计划完成率'] = df[self.year1 + '年累计保养收入'] / df['指标']
        df['同比增长率'] = (df[self.year1 + '年累计保养收入'] / df[self.year2 + '年同期累计收入']) - 1
        df = df[[self.year1 + '年累计保养收入', self.year1 + '年累计免费保养收入', self.year2 + '年同期累计收入', '月度计划完成率', '同比增长率']]

        return df

    def maintenance_profit(self, default):  # 保养利润
        df = self.df2.copy()
        df = df.loc['2.保养']
        df = df.T
        df = df / 10000
        df.columns = [self.year1 + '年累计保养收入', self.year1 + '年累计保养成本']
        df.index.name = '分公司'
        df.reset_index(inplace=True)
        df = self.company_convert(df, default)
        df = df[[(self.year1 + '年累计保养收入'), (self.year1 + '年累计保养成本')]].groupby(df['分公司']).sum()
        df[self.year1 + '年累计保养毛利'] = df[self.year1 + '年累计保养收入'] - df[self.year1 + '年累计保养成本']

        dfly = self.df2ly.copy()
        dfly = dfly.loc['2.保养']
        dfly = dfly.T
        dfly = dfly / 10000
        dfly.columns = [self.year2 + '年同期保养收入', self.year2 + '年同期保养成本']
        dfly.index.name = '分公司'
        dfly.reset_index(inplace=True)
        dfly = self.company_convert(dfly, default)
        dfly = dfly[[(self.year2 + '年同期保养收入'), (self.year2 + '年同期保养成本')]].groupby(dfly['分公司']).sum()
        df[self.year2 + '年同期保养毛利'] = dfly[self.year2 + '年同期保养收入'] - dfly[self.year2 + '年同期保养成本']
        df.loc['合计'] = df.sum()
        df.replace(0, np.nan, inplace=True)
        df[self.year1 + '年累计保养毛利率'] = df[self.year1 + '年累计保养毛利'] / df[self.year1 + '年累计保养收入']
        df.replace(np.nan, 0, inplace=True)
        df = df[[self.year1 + '年累计保养毛利', self.year2 + '年同期保养毛利', self.year1 + '年累计保养毛利率']]

        return df

    def part_income(self, default):  # 备件收入
        df = self.df2.copy()
        df = df.loc[['5.配件', '3.本部配件']]
        df = df.iloc[[0, 2]]
        df = df.T
        df.fillna(0, inplace=True)
        df[self.year1 + '年累计收入'] = df['5.配件'] + df['3.本部配件']
        df = df[[self.year1 + '年累计收入']]
        df.index.name = '分公司'
        df.reset_index(inplace=True)
        df = self.company_convert(df, default)
        df = df[self.year1 + '年累计收入'].groupby(df['分公司']).sum().to_frame()

        dfly = self.df2ly.copy()
        dfly = dfly.loc[['5.配件', '3.本部配件']]
        dfly = dfly.iloc[[0, 2]]
        dfly = dfly.T
        dfly.fillna(0, inplace=True)
        dfly[self.year2 + '年同期累计收入'] = dfly['5.配件'] + dfly['3.本部配件']
        dfly = dfly[[self.year2 + '年同期累计收入']]
        dfly.index.name = '分公司'
        dfly.reset_index(inplace=True)
        dfly = self.company_convert(dfly, default)
        dfly = dfly[self.year2 + '年同期累计收入'].groupby(dfly['分公司']).sum().to_frame()
        df[self.year2 + '年同期累计收入'] = dfly
        df = df / 10000

        df10b = self.df10b.copy()
        df10b = self.company_convert(df10b, default)
        df10b.replace('/', np.nan, inplace=True)
        df10b = df10b['收入'].groupby(df10b['分公司']).sum().to_frame()
        df['指标'] = df10b
        df.loc['合计'] = df.sum()
        df['月度计划完成率'] = df[self.year1 + '年累计收入'] / df['指标']
        df['同比增长率'] = (df[self.year1 + '年累计收入'] / df[self.year2 + '年同期累计收入']) - 1
        del df['指标']

        return df

    def part_profit(self, default):  # 备件毛利
        df = self.df2.copy()
        df = df.loc[['5.配件', '3.本部配件']]
        df = df.T
        df.columns = ['备件收入', '备件成本', '本部备件收入', '本部备件利润']
        df.fillna(0, inplace=True)
        df[self.year1 + '年累计毛利'] = df['备件收入'] - df['备件成本'] + df['本部备件利润']
        df[self.year1 + '年累计备件收入'] = df['备件收入'] + df['本部备件收入']
        df.index.name = '分公司'
        df.reset_index(inplace=True)
        df = self.company_convert(df, default)
        df = df[[self.year1 + '年累计毛利', self.year1 + '年累计备件收入']].groupby(df['分公司']).sum()

        dfly = self.df2ly.copy()
        dfly = dfly.loc[['5.配件', '3.本部配件']]
        dfly = dfly.iloc[[0, 1, 3]]
        dfly = dfly.T
        dfly.columns = ['备件收入', '备件成本', '本部备件利润']
        dfly.fillna(0, inplace=True)
        dfly[self.year2 + '年同期累计毛利'] = dfly['备件收入'] - dfly['备件成本'] + dfly['本部备件利润']
        dfly.index.name = '分公司'
        dfly.reset_index(inplace=True)
        dfly = self.company_convert(dfly, default)
        dfly = dfly[[self.year2 + '年同期累计毛利']].groupby(dfly['分公司']).sum()
        df[self.year2 + '年同期累计毛利'] = dfly
        df.fillna(0, inplace=True)
        df = df / 10000
        df.loc['合计'] = df.sum()
        df['毛利率'] = df[self.year1 + '年累计毛利'] / df[self.year1 + '年累计备件收入']
        del df[self.year1 + '年累计备件收入']
        return df

    def elevator_category(self, default):  # 各梯型占比
        df = self.df1[['数量', '分公司', '产品型号']].copy()
        df = self.company_convert(df, default)
        df = df.groupby(['产品型号']).sum()
        df.index.name = '梯型'
        df.rename(columns={'数量': '台数'}, inplace=True)
        df['占比'] = df['台数'] / df['台数'].sum()
        return df

    def income_percentage(self, default):  # 保养收入占比
        df = self.df2.copy()
        df = df.loc[['服务业收入（分公司报表口径）', '1.安装1', '2.保养']]
        df = df.iloc[[0, 1, 3]]
        df = df.T
        df.columns = ['服务业收入', '安装收入', '保养收入']
        df.index.name = '分公司'
        df.reset_index(inplace=True)
        df = self.company_convert(df, default)
        df = df[['服务业收入', '安装收入', '保养收入']].groupby(df['分公司']).sum()
        df.replace(0, np.nan, inplace=True)
        df['维保收入在工程收入占比'] = 1 - df['安装收入'] / df['服务业收入']
        df['保养收入在维保收入占比'] = df['保养收入'] / (df['服务业收入'] - df['安装收入'])
        df.replace(np.nan, 0, inplace=True)
        df = df[['维保收入在工程收入占比', '保养收入在维保收入占比']]
        return df

    def vip_calculate(self, default):
        df = self.df1[['数量', '分公司', 'VIP名称']].copy()
        df = self.company_convert(df, default)
        df = df.groupby(['VIP名称']).sum()
        df.index.name = '客户名称'
        df.rename(columns={'数量': '总保养台数'}, inplace=True)
        down = df.copy()
        if default == '全国':
            df = df[df['总保养台数'] > 500]
        df.sort_values(by=['总保养台数'], ascending=False, inplace=True)
        if default != '全国':
            df = df.head(n=10)
        df.loc['合计'] = down.sum()
        return df

    def agreement_type(self, default):
        df = self.df1[['数量', '分公司', '实际价/年']].copy()
        df = self.company_convert(df, default)
        dfa = df['数量'].groupby(df['分公司']).sum().to_frame()
        dfa['有偿保养台数'] = df['数量'][df['实际价/年'] > 100].groupby(df['分公司']).sum()
        dfa.replace(np.nan, 0, inplace=True)
        dfa.rename(columns={'数量': '总保养台数'}, inplace=True)
        dfa['免费保养台数'] = dfa['总保养台数'] - dfa['有偿保养台数']
        dfa.loc['合计'] = dfa.sum()
        dfa['免费保养台数占比'] = dfa['免费保养台数'] / dfa['总保养台数']
        df = dfa[['有偿保养台数', '免费保养台数', '免费保养台数占比']]
        return df

    def remake_calculate(self, default):
        df = self.df1[['数量', '分公司', '销售组织']].copy()
        df = self.company_convert(df, default)
        df = df[df['销售组织'].isin(['9大改造梯'])]
        df = df['数量'].groupby(df['分公司']).sum().to_frame()
        df.rename(columns={'数量': '保养台数'}, inplace=True)
        df.loc['合计'] = df.sum()
        return df

    def building_calculate(self, default):
        df = self.df1[['数量', '分公司', '层站门站序', '楼宇用途']].copy()
        df = self.company_convert(df, default)
        df['平均层站数'] = df['层站门站序'].str.split(r'''/''', expand=True)[0]
        df['平均层站数'] = pd.to_numeric(df['平均层站数'], errors='coerce')
        dfa = df['平均层站数'].groupby(df['楼宇用途']).mean()
        dfa = dfa.to_frame()
        dfa['台数'] = df['数量'].groupby(df['楼宇用途']).sum()
        dfa.loc['合计', '平均层站数'] = df['平均层站数'].mean()
        dfa.loc['合计', '台数'] = df['数量'].sum()
        dfa['自保养台数占比'] = dfa['台数'] / dfa.loc['合计', '台数']
        dfa = dfa[['平均层站数', '自保养台数占比']]
        return dfa

    def quantity_year(self, default):  # 投放量
        df = self.df1[['数量', '合同号-梯号', '分公司', '类型']].copy()
        if default != '全国':
            company = self.convert_province(default)
            df = df[df['分公司'].isin(company)]
        df['年份'] = df['合同号-梯号'].apply(lambda x: x[:2])
        df['年份'] = pd.to_numeric(df['年份'], errors='coerce', downcast='integer')
        df['年份'] = df['年份'].apply(self.years)
        df['年份'].fillna('15年以上', inplace=True)
        sort_years = ['0-5年', '5-10年', '10-15年', '15年以上']
        df['年份'] = df['年份'].astype('category')
        df['年份'].cat.set_categories(sort_years, inplace=True)
        dfa = df['数量'].groupby(df['年份']).sum().to_frame()
        dfa.rename(columns={'数量': '总保养台数'}, inplace=True)
        df = df[df['类型'].isin(['ZWV'])]
        dfa['自保养台数'] = df['数量'].groupby(df['年份']).sum()
        dfa = dfa.reindex(index=list(dfa.index) + ['合计'])
        dfa.loc['合计'] = dfa.sum()
        return dfa

    def average_price(self, default):  # 保养均价
        df = self.df1[['分公司', '实际价/年', '类型']].copy()
        df = self.company_convert(df, default)
        df = df[~df['类型'].isin(['ZWV2'])]
        df = df[df['实际价/年'] > 100]
        dfa = df['实际价/年'].groupby(df['分公司']).mean().to_frame()
        dfa.rename(columns={'实际价/年': '保养均价'}, inplace=True)
        dfa['自保养均价'] = df['实际价/年'][df['类型'].isin(['ZWV'])].groupby(df['分公司']).mean()
        dfa['三方合同均价'] = df['实际价/年'][df['类型'].isin(['ZWV5'])].groupby(df['分公司']).mean()
        dfa.loc['合计'] = [df['实际价/年'].mean(), df['实际价/年'][df['类型'].isin(['ZWV'])].mean(),
                         df['实际价/年'][df['类型'].isin(['ZWV5'])].mean()]
        return dfa

    def maintenance_mode(self, default):  # 全包、半包、清包
        df = self.df1[['数量', '分公司', '保养方式']].copy()
        if default != '全国':
            company = self.convert_province(default)
            df = df[df['分公司'].isin(company)]
        df = df['数量'].groupby(df['保养方式']).sum().to_frame()
        df.rename(columns={'数量': '台数'}, inplace=True)
        df.loc['合计'] = df.sum()
        df['占比'] = df['台数'] / df.loc['合计', '台数']
        return df

    def lehy3_calculate(self, default):
        df = self.df1[['数量', '分公司', '产品型号']].copy()
        df = self.company_convert(df, default)
        df = df[df['产品型号'].isin(['LEHY-III'])]
        df = df['数量'].groupby(df['分公司']).sum().to_frame()
        df.rename(columns={'数量': '累计保养'}, inplace=True)
        df.loc['合计'] = df.sum()
        return df

    def mrl_calculate(self, default):
        df = self.df1[['数量', '分公司', '产品型号']].copy()
        df = self.company_convert(df, default)
        df = df[df['产品型号'].isin(['LEHY-MRL'])]
        df = df['数量'].groupby(df['分公司']).sum().to_frame()
        df.rename(columns={'数量': '累计保养'}, inplace=True)
        df.loc['合计'] = df.sum()
        return df

    def mese_calculate(self, default):
        df = self.df1[['数量', '分公司', '产品型号']].copy()
        df = self.company_convert(df, default)
        df = df[df['产品型号'].isin(['MESE'])]
        df = df['数量'].groupby(df['分公司']).sum().to_frame()
        df.rename(columns={'数量': '累计保养'}, inplace=True)
        df.loc['合计'] = df.sum()
        return df

    def site_calculate(self, default):
        df = self.df3.copy()
        df = self.company_convert(df, default)
        df = df['机构名称'].groupby([df['分公司'], df['机构名称']]).count().to_frame(name='数量')
        df.reset_index(level=1, inplace=True)
        df = df.pivot(columns='机构名称', values='数量')
        try:
            df = df[['维保中心', '维保站', '维保点']]  # 注意这个是try的，因为很多不一定有维保点
        except:
            pass
        df['总计'] = df.sum(axis=1)
        df.loc['合计'] = df.sum()
        df.fillna(0, inplace=True)
        return df

    def staff_calculate(self, default):
        df = self.df4.copy()
        if default != '全国':
            company = self.convert_province(default)
            df = df[df['分公司'].isin(company)]
        df.drop(columns=['分公司'], inplace=True)
        df.loc['人数'] = df.sum()
        df['合计'] = df.sum(axis=1)
        df = df.loc['人数'].to_frame()
        df.index.name = '岗位'
        return df

    def remote_calculate(self, default):
        df = self.df5.copy()
        df = self.company_convert(df, default)
        dfa = df['电梯在线状态'].groupby(df['分公司']).count().to_frame()
        dfa.rename(columns={'电梯在线状态': '安装台数'}, inplace=True)
        dfa['在线台数'] = df[df['电梯在线状态'].isin(['在线'])].groupby('分公司').count()
        dfa.loc['合计'] = dfa.sum()
        dfa['在线率'] = dfa['在线台数'] / dfa['安装台数']
        dfa = dfa[['在线台数', '在线率']]
        return dfa

    def urgent_repair(self, default, dfz):
        df = self.df6.copy()
        df = self.company_convert(df, default)
        dfa = (df[df['故障性质'].isin(['3、困人'])])['故障性质'].groupby(df['分公司']).count().to_frame()
        dfa.rename(columns={'故障性质': '困人故障数'}, inplace=True)
        dfa['急修数'] = df['故障性质'].groupby(df['分公司']).count()
        dfa['故障数'] = df[df['非产品故障'].isin([np.nan])]['合同类型'].groupby(df['分公司']).count()
        dfa.loc['合计'] = dfa.sum()
        try:
            dfa['自保养台数'] = dfz['自保养台数']
            dfa['急修率'] = dfa['急修数'] / dfa['自保养台数']
            dfa['故障率'] = dfa['故障数'] / dfa['自保养台数']
            dfa['困人率'] = dfa['困人故障数'] / dfa['自保养台数']
            dfa = dfa[['困人故障数', '急修率', '故障率', '困人率']]
        except:
            pass
        return dfa

    def satisfaction_calculate(self, default):
        df = self.df7.copy()
        df = self.company_convert(df, default)
        df = df.groupby(df['分公司']).mean()
        df.loc['合计'] = df.mean()
        return df

    def maintenance_urgent(self, default, dfz):
        df = self.df8.copy()
        df = self.company_convert(df, default)
        df = df.groupby(df['分公司']).sum()
        dfa = self.df6.copy()
        dfa = dfa[dfa['受信方式'].isin(['电话报修'])]
        df['急修移动终端建单数'] = dfa['受信方式'].groupby(dfa['分公司']).count()
        try:
            df['自保养台数'] = dfz['自保养台数']
        except:
            pass
        df.loc['合计'] = df.sum()
        df['保养移动终端综合执行率'] = df['综合执行工单数'] / df['工单总数']
        try:
            df['急修移动终端建单率'] = df['急修移动终端建单数'] / df['自保养台数']
            df = df[['保养移动终端综合执行率', '急修移动终端建单率']]
        except:
            pass
        return df

    # 二、pandas计算函数（结束）

    # 三、自动出图各函数（开始）
    def maintenance_volume(self, wb, default):
        # 本函数负责完成保养实物量的相关内容填写
        # ws是出图的表，ws1是本来的保养实物量
        ws = wb.create_sheet(title='图表', index=0)  # 因为是第一段内容，所以新建图表
        ws1 = wb['保养实物量']
        self.format_change(ws1, '保养市场占有率', numbers.FORMAT_PERCENTAGE_00)
        self.format_change(ws1, '总保养月度计划完成率', numbers.FORMAT_PERCENTAGE_00)
        self.format_change(ws1, '总保养净增计划完成率', numbers.FORMAT_PERCENTAGE_00)
        ws['A1'] = '◇' + default + '台数指标完成情况'
        ws['B2'] = '总台数'
        ws['B3'].value = self.value_get(ws1, '总保养台数')  # 自保养
        ws['C3'].value = '月度指标完成率'
        ws['C4'].value = self.value_get(ws1, '总保养月度计划完成率')
        ws['C4'].number_format = numbers.FORMAT_PERCENTAGE
        ws['D3'].value = '同比增长'
        ws['E3'].value = '市场占有率'
        ws['E4'].value = self.value_get(ws1, '保养市场占有率')
        ws['E4'].number_format = numbers.FORMAT_PERCENTAGE
        ws['J2'] = '自保养台数'
        ws['O2'] = '三方合同台数'
        ws['R2'] = '三方合作台数'

        ws['B3'].value = self.value_get(ws1, '总保养台数')  # 自保养
        ws['J3'].value = self.value_get(ws1, '自保养台数')  # 三方合同
        ws['K3'].value = '占比'
        ws['K4'].value = ws['J3'].value / ws['B3'].value  # 自保养占比
        ws['K4'].number_format = numbers.FORMAT_PERCENTAGE
        ws['O3'].value = self.value_get(ws1, '三方合同')  # 三方合作
        ws['Q3'].value = '占比'
        ws['Q4'].value = ws['O3'].value / ws['B3'].value  # 三方保养占比
        ws['Q4'].number_format = numbers.FORMAT_PERCENTAGE
        ws['R3'].value = self.value_get(ws1, '三方合作')

        # 占比别忘了到时处理下百分比格式
        width = (7 + 1.2 * 4 * (ws1.max_row - 2))
        data1 = Reference(ws1, min_col=2, min_row=1, max_col=5, max_row=(ws1.max_row - 1))
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row - 1)
        a1 = self.barchart_function(width, data1, cats, '保养实物量', '台数', style1=10)
        data2 = Reference(ws1, min_col=6, min_row=1, max_col=ws1.max_column, max_row=(ws1.max_row - 1))
        a2 = self.linechart_function(data2, '百分比')
        a1.y_axis.crosses = 'max'
        a1 += a2
        ws.add_chart(a1, 'A6')
        return wb

    def income_profit(self, wb, default):
        # 本函数负责完成保养收入到备件毛利率的相关内容填写
        # ws是出图的表，ws1-4是其他几个表
        ws = wb['图表']
        ws1 = wb['保养收入']  # 每个表后面跟进格式修改的内容
        self.format_change(ws1, self.year1 + '年累计保养收入', numbers.FORMAT_NUMBER_00)
        self.format_change(ws1, self.year1 + '年累计免费保养收入', numbers.FORMAT_NUMBER_00)
        self.format_change(ws1, self.year2 + '年同期累计收入', numbers.FORMAT_NUMBER_00)
        self.format_change(ws1, '月度计划完成率', numbers.FORMAT_PERCENTAGE_00)
        self.format_change(ws1, '同比增长率', numbers.FORMAT_PERCENTAGE_00)
        ws2 = wb['保养毛利']
        self.format_change(ws2, self.year1 + '年累计保养毛利', numbers.FORMAT_NUMBER_00)
        self.format_change(ws2, self.year2 + '年同期保养毛利', numbers.FORMAT_NUMBER_00)
        self.format_change(ws2, self.year1 + '年累计保养毛利率', numbers.FORMAT_PERCENTAGE_00)
        ws3 = wb['备件收入']
        self.format_change(ws3, self.year1 + '年累计收入', numbers.FORMAT_NUMBER_00)
        self.format_change(ws3, self.year2 + '年同期累计收入', numbers.FORMAT_NUMBER_00)
        self.format_change(ws3, '月度计划完成率', numbers.FORMAT_PERCENTAGE_00)
        self.format_change(ws3, '同比增长率', numbers.FORMAT_PERCENTAGE_00)
        ws4 = wb['备件毛利']
        self.format_change(ws4, self.year1 + '年累计毛利', numbers.FORMAT_NUMBER_00)
        self.format_change(ws4, self.year2 + '年同期累计毛利', numbers.FORMAT_NUMBER_00)
        self.format_change(ws4, '毛利率', numbers.FORMAT_PERCENTAGE_00)
        ws['A20'] = '◇' + default + '收入指标完成情况（万元）'
        ws['B21'] = '保养收入'
        ws['F21'] = '保养毛利'
        ws['I21'] = '备件收入'
        ws['O21'] = '备件毛利'
        param = self.year1 + '年累计保养收入'
        ws['B22'] = self.value_get(ws1, param)
        ws['D22'] = '同比增长'
        ws['E22'] = '免费保养收入'
        param = self.year1 + '年累计免费保养收入'
        ws['E23'] = self.value_get(ws1, param)
        param = self.year1 + '年累计保养毛利'
        ws['F22'] = self.value_get(ws2, param)
        ws['F22'].number_format = numbers.FORMAT_NUMBER_00
        ws['G22'] = '毛利率'
        param = self.year1 + '年累计保养毛利率'
        ws['G23'] = self.value_get(ws2, param)
        ws['G23'].number_format = numbers.FORMAT_PERCENTAGE_00
        param = self.year1 + '年累计收入'
        ws['I22'] = self.value_get(ws3, param)
        param = self.year1 + '年累计备件毛利'
        ws['O22'] = self.value_get(ws4, param)
        ws['Q22'] = '毛利率'
        param = self.year1 + '年累计备件毛利率'
        ws['Q23'] = self.value_get(ws4, param)
        # 第一个表，保养收入
        width = (7 + 1.2 * 3 * (ws1.max_row - 2))
        data1 = Reference(ws1, min_col=2, min_row=1, max_col=4, max_row=(ws1.max_row - 1))
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row - 1)
        a1 = self.barchart_function(width, data1, cats, '保养收入（万元）', '收入', style1=30)
        data2 = Reference(ws1, min_col=5, max_col=ws1.max_column, min_row=1, max_row=(ws1.max_row - 1))
        a2 = self.linechart_function(data2, '百分比')
        a1 += a2
        a1.y_axis.crosses = 'max'
        ws.add_chart(a1, 'A24')
        # 第二个表 保养毛利
        width = (7 + 1.2 * 2 * (ws2.max_row - 2))
        data1 = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=(ws2.max_row - 1))
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row - 1)
        b1 = self.barchart_function(width, data1, cats, '保养毛利（万元）', '毛利润', style1=4)
        data2 = Reference(ws2, min_col=ws2.max_column, min_row=1, max_row=(ws2.max_row - 1))
        b2 = self.linechart_function(data2, '毛利率')
        b2s1 = b2.series[0]
        b2s1.marker.symbol = 'circle'
        b1.y_axis.crosses = 'max'
        # b2条形图准备完毕
        b1 += b2
        ws.add_chart(b1, 'A39')
        # 第三个表，备件收入
        width = (7 + 1.2 * 2 * (ws3.max_row - 2))
        data1 = Reference(ws3, min_col=2, max_col=3, min_row=1, max_row=(ws3.max_row - 1))
        cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row - 1)
        c1 = self.barchart_function(width, data1, cats, '备件收入（万元）', '收入')
        data2 = Reference(ws3, min_col=4, max_col=ws3.max_column, min_row=1, max_row=(ws3.max_row - 1))
        c2 = self.linechart_function(data2, '百分比')
        c1 += c2
        c1.y_axis.crosses = 'max'
        ws.add_chart(c1, 'A54')
        # 第四个表，备件利润
        width = (7 + 1.2 * 2 * (ws4.max_row - 2))
        data1 = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=(ws4.max_row - 1))
        cats = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row - 1)
        d1 = self.barchart_function(width, data1, cats, '备件毛利（万元）', '毛利润', style1=5)
        # d1条形图准备完毕
        data2 = Reference(ws4, min_col=ws4.max_column, min_row=1, max_row=(ws4.max_row - 1))
        d2 = self.linechart_function(data2, '毛利率')
        d2s1 = d2.series[0]
        d2s1.marker.symbol = 'circle'
        d1.y_axis.crosses = 'max'
        d1 += d2
        ws.add_chart(d1, 'A69')
        return wb

    def incomepercentage_write(self, wb):
        ws = wb['图表']
        ws1 = wb['收入占比']
        self.format_change(ws1, '维保收入在工程收入占比', numbers.FORMAT_PERCENTAGE_00)
        self.format_change(ws1, '保养收入在维保收入占比', numbers.FORMAT_PERCENTAGE_00)
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
        data = Reference(ws1, min_col=2, min_row=1, max_col=ws1.max_column, max_row=ws1.max_row)
        chart = self.radarchart_function(data, cats, '保养收入占比', style1=26)
        ws.add_chart(chart, "A84")
        return wb

    def elevator_write(self, wb):
        ws = wb['图表']
        ws1 = wb['梯型分布']
        self.format_change(ws1, '占比', numbers.FORMAT_PERCENTAGE_00)
        ws['A127'] = '在保养电梯梯型分布'
        for i in range(1, ws1.max_row + 1):
            for j in range(1, ws1.max_column + 1):
                ws['A128'].offset((i - 1), (j - 1)).value = ws1.cell(i, j).value
                ws['A128'].offset((i - 1), (j - 1)).number_format = ws1.cell(i, j).number_format
        return wb

    @staticmethod
    def vip_write(wb, default):
        ws = wb['图表']
        ws1 = wb['战略客户']
        if default == '全国':
            ws['F127'].value = '保养台数大于500台战略客户列表'
        else:
            ws['F127'].value = '保养台数前10战略客户列表'
        for i in range(1, ws1.max_row + 1):
            for j in range(1, ws1.max_column + 1):
                ws['F128'].offset((i - 1), (j - 1)).value = ws1.cell(i, j).value
        return wb

    def agreement_analyze(self, wb):
        ws = wb['图表']
        ws2 = wb['保养合同类型']
        self.format_change(ws2, '免费保养台数占比', numbers.FORMAT_PERCENTAGE_00)
        ws4 = wb['旧梯改造']
        ws['A155'].value = '◇　保养合同分析'
        ws['B156'].value = '保养流失新签'
        ws['F156'].value = '保养合同类型'
        ws['I156'].value = '旧梯改造保养受控率'
        ws['M156'].value = '自保养免保合同转签'
        ws['P156'].value = '三方合同免费保养转签'
        ws['B157'].value = '新签'
        ws['C157'].value = '累计流失'
        ws['D157'].value = '累计流失率'
        ws['E157'].value = '到期流失率'
        # 保养合同类型
        ws['F157'].value = '有偿保养'
        ws['F158'].value = self.value_get(ws2, '有偿保养台数')
        ws['G157'].value = '免费保养'
        ws['G158'].value = self.value_get(ws2, '免费保养台数')
        ws['H157'].value = '免保占比'
        ws['H158'].value = self.value_get(ws2, '免费保养台数占比')
        ws['H158'].number_format = numbers.FORMAT_PERCENTAGE_00
        # 旧梯改造保养受控率
        ws['I157'].value = '销售台数'
        ws['J157'].value = '安装台数'
        ws['K157'].value = '保养台数'
        ws['K158'].value = self.value_get(ws4, '保养台数')
        ws['L157'].value = '受控率'
        ws['M157'].value = '到期台数'
        ws['N157'].value = '到期转签率'
        ws['P157'].value = '到期台数'
        ws['R157'].value = '到期转签率'

        # 图2
        width = 7 + 1.2 * 2 * (ws2.max_row - 2)
        data1 = Reference(ws2, min_col=2, max_col=(ws2.max_column - 1), min_row=1, max_row=(ws2.max_row - 1))
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row - 1)
        b1 = self.barchart_function(width, data1, cats, '保养合同类型', '台数', style1=7)
        data2 = Reference(ws2, min_col=ws2.max_column, min_row=1, max_row=(ws2.max_row - 1))
        b2 = self.linechart_function(data2, '占比')
        b2s1 = b2.series[0]
        b2s1.marker.symbol = 'star'
        b1.y_axis.crosses = 'max'
        b1 += b2
        ws.add_chart(b1, 'A174')
        # 试试看不设定颜色

        # 图4
        width = 4.74 + 1.2 * 1 * (ws4.max_row - 2)
        data1 = Reference(ws4, min_col=2, min_row=1, max_row=(ws4.max_row - 1))
        cats = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row - 1)
        d1 = self.barchart_function(width, data1, cats, '旧梯改造保养受控率', '台数', style1=31)
        ws.add_chart(d1, 'A203')

        return wb

    def building_write(self, wb):
        ws = wb['图表']
        ws1 = wb['楼宇类型']
        self.format_change(ws1, '平均层站数', numbers.FORMAT_NUMBER)
        self.format_change(ws1, '自保养台数占比', numbers.FORMAT_PERCENTAGE_00)
        width = 7 + 1.2 * 1 * (ws1.max_row - 2)
        data1 = Reference(ws1, min_col=2, min_row=1, max_row=(ws1.max_row - 1))
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row - 1)
        a1 = self.barchart_function(width, data1, cats, '客户类型占比及平均层站数', '层站数', style1=15)
        data2 = Reference(ws1, min_col=ws1.max_column, min_row=1, max_row=(ws1.max_row - 1))
        a2 = self.linechart_function(data2, '占比')
        a2s1 = a2.series[0]
        a2s1.marker.symbol = 'picture'
        a1.y_axis.crosses = 'max'
        a1 += a2
        ws.add_chart(a1, 'A218')
        return wb

    def quantity_write(self, wb):  # 这个函数出投放量的图
        ws = wb['图表']
        ws1 = wb['投放量']
        width = 30
        data1 = Reference(ws1, min_col=2, min_row=1, max_row=(ws1.max_row - 1), max_col=ws1.max_column)
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row - 1)
        a1 = self.areachart_function(width, data1, cats, '投放年限保养占比分析')
        ws.add_chart(a1, 'A233')
        return wb

    def price_analyze(self, wb):
        ws = wb['图表']
        ws1 = wb['保养平均单价']
        self.format_change(ws1, '保养均价', numbers.FORMAT_NUMBER)
        self.format_change(ws1, '自保养均价', numbers.FORMAT_NUMBER)
        self.format_change(ws1, '三方合同均价', numbers.FORMAT_NUMBER)
        ws3 = wb['保养方式']
        self.format_change(ws3, '占比', numbers.FORMAT_PERCENTAGE_00)
        ws['A248'].value = '◇ 保养价格分析'
        ws['B249'].value = '保养合同均价（年/台/元）'
        ws['B250'].value = self.value_get(ws1, '保养均价')
        ws['B250'].number_format = numbers.FORMAT_NUMBER_00
        ws['C250'].value = '自保养均价'
        ws['C251'].value = self.value_get(ws1, '自保养均价')
        ws['C251'].number_format = numbers.FORMAT_NUMBER_00
        ws['D250'].value = '三方保养均价'
        ws['D251'].value = self.value_get(ws1, '三方合同均价')
        ws['D251'].number_format = numbers.FORMAT_NUMBER_00
        ws['E249'].value = '销售预留免保单价（年/台/元）'
        ws['E250'].value = '预留免保单价'
        ws['F250'].value = '台数（18编号安装合同）'
        ws['I249'].value = '保养合同方式（台)'
        ws['I250'].value = '全包'
        param = '台数'
        ws['I251'].value = self.value_get(ws3, param, '全包')
        ws['J250'].value = '半包'
        ws['J251'].value = self.value_get(ws3, param, '半包')
        ws['K250'].value = '清包'
        ws['K251'].value = self.value_get(ws3, param, '清包')
        ws['N250'].value = '非标'
        ws['N251'].value = self.value_get(ws3, param, '非标')
        width = (4.74 + 1.2 * 1.6 * (ws1.max_row - 2))
        data1 = Reference(ws1, min_col=2, min_row=1, max_row=ws1.max_row - 1, max_col=ws1.max_column)
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row - 1)
        a1 = self.barchart_function(width, data1, cats, '保养合同均价', '价格', style1=15, grouping1='stacked')
        ws.add_chart(a1, 'A252')

        data3 = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row - 1)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row - 1)
        slices = [DataPoint(idx=i) for i in range(20)]
        c1 = DoughnutChart()
        c1.add_data(data3, titles_from_data=True)
        c1.set_categories(cats)
        c1.data_points = slices
        c1.title = '保养合同方式情况'
        c1.dLbls = DataLabelList()
        c1.dLbls.showVal = True
        c1.dLbls.showCatName = True
        c1.dLbls.showPercent = True
        c1.dLbls.position = 'bestFit'
        ws.add_chart(c1, 'A282')

        return wb

    def special_percentage(self, wb):
        ws = wb['图表']
        ws1 = wb['LEHY-3保养受控率']
        ws2 = wb['LEHY-MRL保养受控率']
        ws3 = wb['MESE保养受控率']
        ws['A297'].value = '◇ 特殊梯种保养受控率'
        ws['B298'].value = 'LEHY-III'
        ws['B299'].value = '销售台数'
        ws['C299'].value = '安装台数'
        ws['D299'].value = '保养台数'
        ws['D300'].value = self.value_get(ws1, '累计保养')
        ws['E299'].value = '保养受控率'
        ws['G298'].value = 'LEHY_MRL'
        ws['G299'].value = '销售台数'
        ws['H299'].value = '安装台数'
        ws['I299'].value = '保养台数'
        ws['I300'].value = self.value_get(ws2, '累计保养')
        ws['J299'].value = '保养受控率'
        ws['L298'].value = 'MESE'
        ws['L299'].value = '销售台数'
        ws['N299'].value = '安装台数'
        ws['P299'].value = '保养台数'
        ws['P300'].value = self.value_get(ws3, '累计保养')
        ws['R299'].value = '保养受控率'
        width = 4.74 + 1.2 * 1.5 * (ws1.max_row - 2)
        data1 = Reference(ws1, min_col=2, min_row=1, max_col=ws1.max_column, max_row=(ws1.max_row - 1))
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row - 1)
        a1 = self.barchart_function(width, data1, cats, 'LEHY-III保养受控率', '台数')
        ws.add_chart(a1, 'A301')
        width = 4.74 + 1.2 * 1.5 * (ws2.max_row - 2)
        data2 = Reference(ws2, min_col=2, min_row=1, max_col=ws2.max_column, max_row=(ws2.max_row - 1))
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row - 1)
        b1 = self.barchart_function(width, data2, cats, 'LEHY-MRL保养受控率', '台数')
        ws.add_chart(b1, 'A316')
        width = 4.74 + 1.2 * 1.5 * (ws3.max_row - 2)
        data3 = Reference(ws3, min_col=2, min_row=1, max_col=ws3.max_column, max_row=(ws3.max_row - 1))
        cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row - 1)
        c1 = self.barchart_function(width, data3, cats, 'MESE电梯保养受控率', '台数')
        ws.add_chart(c1, 'A331')
        return wb

    def resource_write(self, wb):
        ws = wb['图表']
        ws1 = wb['维保网络']
        ws2 = wb['维保人员']
        ws['A346'].value = '◇ 维保资源'
        ws['B347'].value = '维保站点（个）'
        ws['B348'].value = self.value_get(ws1, '总计')
        ws['C348'].value = '维保中心'
        ws['C349'].value = self.value_get(ws1, '维保中心')
        ws['D348'].value = '维保站'
        ws['D349'].value = self.value_get(ws1, '维保站')
        ws['E348'].value = '维保点'
        ws['E349'].value = self.value_get(ws1, '维保点')
        ws['G347'].value = '维保人力资源'
        ws['G348'].value = self.value_get(ws2, '人数')
        ws['H348'].value = '新入职人数'
        ws['J348'].value = '今年离职人数'
        ws['M348'].value = '离职率'
        width = 4.74 + 1.2 * 3 * (ws1.max_row - 2)
        data1 = Reference(ws1, min_col=2, min_row=1, max_col=ws1.max_column, max_row=(ws1.max_row - 1))
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row - 1)
        a1 = self.barchart_function(width, data1, cats, '维保站点小计', '数量')
        ws.add_chart(a1, 'A350')
        b1 = ProjectedPieChart()
        data2 = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row - 1, max_col=ws2.max_column)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row - 1)
        b1.add_data(data2, titles_from_data=True)
        b1.set_categories(cats)
        b1.dLbls = DataLabelList()
        b1.dLbls.showVal = True
        b1.dLbls.showCatName = True
        b1.dLbls.position = 'bestFit'
        ws.add_chart(b1, 'A365')
        return wb

    def last_write(self, wb):
        ws = wb['图表']
        ws1 = wb['远程监控台数']
        self.format_change(ws1, '在线率', numbers.FORMAT_PERCENTAGE_00)
        ws2 = wb['急修情况分析']
        try:
            self.format_change(ws2, '急修率', numbers.FORMAT_PERCENTAGE_00)
            self.format_change(ws2, '故障率', numbers.FORMAT_PERCENTAGE_00)
            self.format_change(ws2, '困人率', numbers.FORMAT_PERCENTAGE_00)
        except:
            pass
        ws3 = wb['用户满意度']
        self.format_change(ws3, '保养用户满意度评分', numbers.FORMAT_NUMBER_00)
        self.format_change(ws3, '备件供应服务', numbers.FORMAT_NUMBER_00)
        ws4 = wb['移动终端']
        self.format_change(ws4, '保养移动终端综合执行率', numbers.FORMAT_PERCENTAGE_00)
        try:
            self.format_change(ws4, '急修移动终端建单率', numbers.FORMAT_PERCENTAGE_00)
        except:
            pass
        ws['B380'].value = '远程监视'
        ws['B381'].value = '在线台数'
        ws['B382'].value = self.value_get(ws1, '在线台数')
        ws['D381'].value = '在线率'
        ws['D382'].value = self.value_get(ws1, '在线率')
        ws['D382'].number_format = numbers.FORMAT_PERCENTAGE_00
        ws['E380'].value = '急修情况'
        ws['E381'].value = '困人故障数'
        ws['E382'].value = self.value_get(ws2, '困人故障数')
        ws['F381'].value = '故障率'
        ws['G381'].value = '急修率'
        ws['H381'].value = '困人率'
        try:
            ws['F382'].value = self.value_get(ws2, '故障率')
            ws['F382'].number_format = numbers.FORMAT_PERCENTAGE_00
            ws['G382'].value = self.value_get(ws2, '急修率')
            ws['G382'].number_format = numbers.FORMAT_PERCENTAGE_00
            ws['H382'].value = self.value_get(ws2, '困人率')
            ws['H382'].number_format = numbers.FORMAT_PERCENTAGE_00
        except:
            pass
        ws['I380'].value = '用户满意度'
        ws['I381'].value = '保养（备件）用户满意度评分'
        ws['I382'].value = self.value_get(ws3, '保养用户满意度评分')
        ws['I382'].number_format = numbers.FORMAT_NUMBER_00
        ws['K381'].value = '备件供应服务'
        ws['K382'].value = self.value_get(ws3, '备件供应服务')
        ws['K382'].number_format = numbers.FORMAT_NUMBER_00
        ws['N380'].value = '维保移动终端'
        ws['N381'].value = '保养终端执行率'
        ws['N382'].value = self.value_get(ws4, '保养移动终端综合执行率')
        ws['N382'].number_format = numbers.FORMAT_PERCENTAGE_00
        ws['Q381'].value = '急修终端建单率'
        ws['Q382'].value = self.value_get(ws4, '急修移动终端建单率')
        ws['Q382'].number_format = numbers.FORMAT_PERCENTAGE_00
        # 第一张图
        width = 7 + 1.2 * 1.2 * (ws1.max_row - 2)
        data1 = Reference(ws1, min_col=2, max_col=(ws1.max_column - 1), min_row=1, max_row=(ws1.max_row - 1))
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row - 1)
        a1 = self.barchart_function(width, data1, cats, '远程监控台数', '台数', style1=7)
        data2 = Reference(ws1, min_col=ws1.max_column, min_row=1, max_row=(ws1.max_row - 1))
        a2 = self.linechart_function(data2, '在线率')
        a2s1 = a2.series[0]
        a2s1.marker.symbol = 'auto'
        a1.y_axis.crosses = 'max'
        a1 += a2
        ws.add_chart(a1, 'A383')
        # 第二张图
        width = 7 + 1.2 * 1.2 * (ws1.max_row - 2)
        data1 = Reference(ws2, min_col=2, min_row=1, max_row=(ws2.max_row - 1))
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row - 1)
        b1 = self.barchart_function(width, data1, cats, '急修情况分析', '台数', style1=10)
        try:
            data2 = Reference(ws2, min_col=3, max_col=ws2.max_column, min_row=1, max_row=(ws2.max_row - 1))
            b2 = self.linechart_function(data2, '百分比')
            b1 += b2
            b1.y_axis.crosses = 'max'
        except:
            pass
        ws.add_chart(b1, 'A398')

        # 第三张图
        width = 4.74 + 1.2 * 2 * (ws3.max_row - 2)
        data1 = Reference(ws3, min_col=2, min_row=1, max_col=ws3.max_column, max_row=(ws3.max_row - 1))
        cats = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row - 1)
        c1 = self.barchart_function(width, data1, cats, '用户满意度情况', '分数', style1=5)
        ws.add_chart(c1, 'A413')

        # 第四张图

        data = Reference(ws4, min_col=2, min_row=1, max_col=ws4.max_column, max_row=(ws4.max_row - 1))
        cats = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row - 1)
        d1 = self.radarchart_function(data, cats, '维保移动终端', style1=5)
        ws.add_chart(d1, 'A428')
        return wb

    # 三、自动出图的函数（结束）

    def op(self, file, default):  # 此函数为出图的总函数
        wb = openpyxl.load_workbook(file)
        wb = self.maintenance_volume(wb, default)
        wb = self.income_profit(wb, default)
        wb = self.elevator_write(wb)
        wb = self.incomepercentage_write(wb)
        wb = self.elevator_write(wb)
        wb = self.vip_write(wb, default)
        wb = self.agreement_analyze(wb)
        wb = self.building_write(wb)
        wb = self.quantity_write(wb)
        wb = self.price_analyze(wb)
        wb = self.special_percentage(wb)
        wb = self.resource_write(wb)
        wb = self.last_write(wb)
        wb.save(file)

        # 哪怕能用，你最后肯定也是要把上面图表变成函数的

    def excel(self, default):  # 这个是专门负责pandas生成数据的函数
        name = '\\各分公司季度报表\\' + default + '季度报表-自动生成.xlsx'
        path = os.getcwd()
        file = path + name
        print(default + '分公司数据计算中....请耐心等待')
        a1 = time.process_time()
        with pd.ExcelWriter(file) as writer:
            dfz = self.maintenance(default)
            dfz.to_excel(writer, '保养实物量')
            (self.maintenance_value(default)).to_excel(writer, '保养收入')
            (self.maintenance_profit(default)).to_excel(writer, '保养毛利')
            (self.part_income(default)).to_excel(writer, '备件收入')
            (self.part_profit(default)).to_excel(writer, '备件毛利')
            (self.elevator_category(default)).to_excel(writer, '梯型分布')
            (self.income_percentage(default)).to_excel(writer, '收入占比')
            (self.vip_calculate(default)).to_excel(writer, '战略客户')
            (self.agreement_type(default)).to_excel(writer, '保养合同类型')
            (self.remake_calculate(default)).to_excel(writer, '旧梯改造')
            (self.building_calculate(default)).to_excel(writer, '楼宇类型')
            (self.quantity_year(default)).to_excel(writer, '投放量')
            (self.average_price(default)).to_excel(writer, '保养平均单价')
            (self.maintenance_mode(default)).to_excel(writer, '保养方式')
            (self.lehy3_calculate(default)).to_excel(writer, 'LEHY-3保养受控率')
            (self.mrl_calculate(default)).to_excel(writer, 'LEHY-MRL保养受控率')
            (self.mese_calculate(default)).to_excel(writer, 'MESE保养受控率')
            (self.site_calculate(default)).to_excel(writer, '维保网络')
            (self.staff_calculate(default)).to_excel(writer, '维保人员')
            (self.remote_calculate(default)).to_excel(writer, '远程监控台数')
            (self.urgent_repair(default, dfz)).to_excel(writer, '急修情况分析')
            (self.satisfaction_calculate(default)).to_excel(writer, '用户满意度')
            (self.maintenance_urgent(default, dfz)).to_excel(writer, '移动终端')

        a2 = time.process_time()
        a = a2 - a1
        print(default + '数据计算完成，耗时%.2f秒' % a)
        print(default + '正在生成图表，请耐心等待')
        b1 = time.process_time()
        self.op(file, default)
        b2 = time.process_time()
        b = b2 - b1
        print(default + '图表生成完成，耗时%.2f秒\n' % b)

    def run(self):
        content = self.list_generate()
        default = '全国'
        self.excel(default)  # 啥都不传就是全国
        for each in content:
            default = each
            self.excel(default)


if __name__ == '__main__':
    print('大家好，今天由王君文给大家带来自制的双月报表自动化程序初版\n')
    print('在此由衷感谢大家对我制作程序的支持，愿我们的心血能提高维保部的办公效率\n')
    print('那么，好戏开场了^_^\n')
    t1 = time.process_time()
    AutoGenerate().run()
    t2 = time.process_time()
    t = t2 - t1
    print('王君文为维保管理科制作的自动化报表已生成，请过目^_^, 共耗时%.2f秒' % t)
